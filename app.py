import os
from datetime import date, datetime, timedelta
from flask import Flask, render_template, request, redirect, url_for, flash, send_file
from models import db, Angajat, Concediu, SarbatoareLegala
from utils import zile_lucratoare, sold_co, sarbatori_legale, get_sarbatori_set

app = Flask(__name__)
app.secret_key = os.environ.get('FLASK_SECRET_KEY', 'concedii-app-secret-2026')

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, 'data', 'concedii.db')
os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{DB_PATH}'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db.init_app(app)

with app.app_context():
    db.create_all()
    # Populare sarbatori legale pt anul curent si urmator
    an_curent = date.today().year
    for an in [an_curent, an_curent + 1]:
        existing = SarbatoareLegala.query.filter_by(an=an).count()
        if existing == 0:
            for d, denumire in sarbatori_legale(an):
                db.session.add(SarbatoareLegala(data=d, denumire=denumire, an=an))
            db.session.commit()

TIPURI_CONCEDIU = {
    'CO': 'Concediu de odihna',
    'MEDICAL': 'Concediu medical',
    'FARA_PLATA': 'Concediu fara plata',
    'EVENIMENT': 'Eveniment (casatorie/deces/nastere)',
}


# ============ DASHBOARD ============

@app.route('/')
def dashboard():
    azi = date.today()
    an = azi.year
    angajati = Angajat.query.filter_by(activ=True).order_by(Angajat.nume).all()

    # Cine e in concediu azi
    in_concediu = Concediu.query.filter(
        Concediu.data_start <= azi, Concediu.data_sfarsit >= azi
    ).all()

    # Sold per angajat
    solduri = []
    for a in angajati:
        concedii = Concediu.query.filter_by(angajat_id=a.id).all()
        co_consumat = sum(c.zile_lucratoare for c in concedii if c.tip == 'CO' and c.data_start.year == an)
        solduri.append({
            'angajat': a,
            'total': a.zile_co_an,
            'consumat': co_consumat,
            'ramas': a.zile_co_an - co_consumat,
        })

    return render_template('dashboard.html',
        angajati=angajati, in_concediu=in_concediu,
        solduri=solduri, azi=azi, an=an, tipuri=TIPURI_CONCEDIU)


# ============ ANGAJATI ============

@app.route('/angajati')
def lista_angajati():
    angajati = Angajat.query.order_by(Angajat.activ.desc(), Angajat.nume).all()
    return render_template('angajati.html', angajati=angajati)


@app.route('/angajati/adauga', methods=['POST'])
def adauga_angajat():
    nume = request.form.get('nume', '').strip()
    prenume = request.form.get('prenume', '').strip()
    departament = request.form.get('departament', '').strip()
    zile_co = request.form.get('zile_co_an', '21')
    data_ang = request.form.get('data_angajare', '')

    if not nume or not prenume:
        flash('Numele si prenumele sunt obligatorii!', 'error')
        return redirect(url_for('lista_angajati'))

    a = Angajat(
        nume=nume, prenume=prenume, departament=departament,
        zile_co_an=int(zile_co) if zile_co else 21,
        data_angajare=datetime.strptime(data_ang, '%Y-%m-%d').date() if data_ang else None,
    )
    db.session.add(a)
    db.session.commit()
    flash(f'Angajat {a.nume_complet} adaugat.', 'success')
    return redirect(url_for('lista_angajati'))


@app.route('/angajati/editeaza/<int:id>', methods=['POST'])
def editeaza_angajat(id):
    a = Angajat.query.get_or_404(id)
    a.nume = request.form.get('nume', a.nume).strip()
    a.prenume = request.form.get('prenume', a.prenume).strip()
    a.departament = request.form.get('departament', a.departament).strip()
    zile_co = request.form.get('zile_co_an', '')
    if zile_co:
        a.zile_co_an = int(zile_co)
    data_ang = request.form.get('data_angajare', '')
    if data_ang:
        a.data_angajare = datetime.strptime(data_ang, '%Y-%m-%d').date()
    activ = request.form.get('activ')
    a.activ = activ == 'on'
    db.session.commit()
    flash(f'Angajat {a.nume_complet} actualizat.', 'success')
    return redirect(url_for('lista_angajati'))


@app.route('/angajati/sterge/<int:id>', methods=['POST'])
def sterge_angajat(id):
    a = Angajat.query.get_or_404(id)
    db.session.delete(a)
    db.session.commit()
    flash(f'Angajat {a.nume_complet} sters.', 'success')
    return redirect(url_for('lista_angajati'))


# ============ CONCEDII ============

@app.route('/concedii')
def lista_concedii():
    an = request.args.get('an', date.today().year, type=int)
    angajat_id = request.args.get('angajat_id', 0, type=int)

    q = Concediu.query.join(Angajat)
    if angajat_id:
        q = q.filter(Concediu.angajat_id == angajat_id)
    concedii = q.filter(
        db.extract('year', Concediu.data_start) == an
    ).order_by(Concediu.data_start.desc()).all()

    angajati = Angajat.query.filter_by(activ=True).order_by(Angajat.nume).all()
    return render_template('concedii.html',
        concedii=concedii, angajati=angajati, an=an,
        angajat_id=angajat_id, tipuri=TIPURI_CONCEDIU)


@app.route('/concedii/adauga', methods=['POST'])
def adauga_concediu():
    angajat_id = request.form.get('angajat_id', type=int)
    tip = request.form.get('tip', 'CO')
    data_start = request.form.get('data_start', '')
    data_sfarsit = request.form.get('data_sfarsit', '')
    observatii = request.form.get('observatii', '').strip()

    if not angajat_id or not data_start or not data_sfarsit:
        flash('Toate campurile sunt obligatorii!', 'error')
        return redirect(url_for('lista_concedii'))

    ds = datetime.strptime(data_start, '%Y-%m-%d').date()
    de = datetime.strptime(data_sfarsit, '%Y-%m-%d').date()

    if ds > de:
        flash('Data start nu poate fi dupa data sfarsit!', 'error')
        return redirect(url_for('lista_concedii'))

    zile = zile_lucratoare(ds, de)

    c = Concediu(
        angajat_id=angajat_id, data_start=ds, data_sfarsit=de,
        tip=tip, zile_lucratoare=zile, observatii=observatii,
    )
    db.session.add(c)
    db.session.commit()
    flash(f'Concediu adaugat: {zile} zile lucratoare.', 'success')
    return redirect(url_for('lista_concedii'))


@app.route('/concedii/sterge/<int:id>', methods=['POST'])
def sterge_concediu(id):
    c = Concediu.query.get_or_404(id)
    db.session.delete(c)
    db.session.commit()
    flash('Concediu sters.', 'success')
    return redirect(url_for('lista_concedii'))


# ============ CALENDAR ============

@app.route('/calendar')
def calendar_view():
    an = request.args.get('an', date.today().year, type=int)
    luna = request.args.get('luna', date.today().month, type=int)

    # Sarbatori legale
    sarbatori = get_sarbatori_set(an)
    sarbatori_dict = {s[0]: s[1] for s in sarbatori_legale(an)}

    # Concedii in luna
    prima_zi = date(an, luna, 1)
    if luna == 12:
        ultima_zi = date(an, 12, 31)
    else:
        ultima_zi = date(an, luna + 1, 1) - timedelta(days=1)

    concedii = Concediu.query.join(Angajat).filter(
        Concediu.data_start <= ultima_zi,
        Concediu.data_sfarsit >= prima_zi,
        Angajat.activ == True,
    ).all()

    # Build calendar data
    zile = []
    d = prima_zi
    while d <= ultima_zi:
        zi_info = {
            'data': d,
            'weekend': d.weekday() >= 5,
            'sarbatoare': sarbatori_dict.get(d, ''),
            'concedii': [],
        }
        for c in concedii:
            if c.data_start <= d <= c.data_sfarsit:
                zi_info['concedii'].append(c)
        zile.append(zi_info)
        d += timedelta(days=1)

    luni_nume = ['', 'Ianuarie', 'Februarie', 'Martie', 'Aprilie', 'Mai', 'Iunie',
                 'Iulie', 'August', 'Septembrie', 'Octombrie', 'Noiembrie', 'Decembrie']

    return render_template('calendar.html',
        zile=zile, an=an, luna=luna, luna_nume=luni_nume[luna],
        tipuri=TIPURI_CONCEDIU)


# ============ SARBATORI ============

@app.route('/sarbatori')
def lista_sarbatori():
    an = request.args.get('an', date.today().year, type=int)
    sarbatori = SarbatoareLegala.query.filter_by(an=an).order_by(SarbatoareLegala.data).all()
    return render_template('sarbatori.html', sarbatori=sarbatori, an=an)


# ============ RAPOARTE ============

@app.route('/rapoarte/export')
def export_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    import tempfile

    an = request.args.get('an', date.today().year, type=int)
    angajati = Angajat.query.filter_by(activ=True).order_by(Angajat.nume).all()

    wb = Workbook()
    ws = wb.active
    ws.title = f'Concedii {an}'

    # Header
    headers = ['Nr.', 'Nume', 'Prenume', 'Departament', 'Zile CO/an',
               'CO consumat', 'CO ramas', 'Medical', 'Fara plata', 'Eveniment', 'Total zile']
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'))

    for j, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=j, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    for i, a in enumerate(angajati, 1):
        concedii = Concediu.query.filter_by(angajat_id=a.id).filter(
            db.extract('year', Concediu.data_start) == an).all()
        co = sum(c.zile_lucratoare for c in concedii if c.tip == 'CO')
        med = sum(c.zile_lucratoare for c in concedii if c.tip == 'MEDICAL')
        fp = sum(c.zile_lucratoare for c in concedii if c.tip == 'FARA_PLATA')
        ev = sum(c.zile_lucratoare for c in concedii if c.tip == 'EVENIMENT')
        vals = [i, a.nume, a.prenume, a.departament, a.zile_co_an,
                co, a.zile_co_an - co, med, fp, ev, co + med + fp + ev]
        for j, v in enumerate(vals, 1):
            cell = ws.cell(row=i + 1, column=j, value=v)
            cell.border = thin_border
            if j >= 5:
                cell.alignment = Alignment(horizontal='center')

    # Auto width
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

    filepath = os.path.join(tempfile.gettempdir(), f'Concedii_{an}.xlsx')
    wb.save(filepath)
    return send_file(filepath, as_attachment=True, download_name=f'Raport Concedii {an}.xlsx')


# ============ API NOTIFICARI ============

@app.route('/api/notifications')
def api_notifications():
    """Returneaza notificari: cine incepe/termina concediu maine, sold scazut."""
    from flask import jsonify
    azi = date.today()
    maine = azi + timedelta(days=1)
    notificari = []

    # Concedii care incep maine
    start_maine = Concediu.query.join(Angajat).filter(
        Concediu.data_start == maine, Angajat.activ == True
    ).all()
    for c in start_maine:
        notificari.append({
            'title': 'Concediu incepe maine',
            'body': f'{c.angajat.nume_complet} - {TIPURI_CONCEDIU.get(c.tip, c.tip)} ({c.zile_lucratoare} zile)',
            'type': 'start',
        })

    # Concedii care se termina azi (revin maine)
    sfarsit_azi = Concediu.query.join(Angajat).filter(
        Concediu.data_sfarsit == azi, Angajat.activ == True
    ).all()
    for c in sfarsit_azi:
        notificari.append({
            'title': 'Revine din concediu',
            'body': f'{c.angajat.nume_complet} revine maine la lucru',
            'type': 'return',
        })

    # Sold CO scazut (sub 5 zile)
    an = azi.year
    angajati = Angajat.query.filter_by(activ=True).all()
    for a in angajati:
        co_consumat = sum(
            c.zile_lucratoare for c in Concediu.query.filter_by(angajat_id=a.id, tip='CO').filter(
                db.extract('year', Concediu.data_start) == an).all()
        )
        ramas = a.zile_co_an - co_consumat
        if 0 < ramas <= 3:
            notificari.append({
                'title': 'Sold CO scazut',
                'body': f'{a.nume_complet} mai are doar {ramas} zile CO ramase',
                'type': 'warning',
            })

    return jsonify(notificari)


# ============ RUN ============

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5080, debug=False)
